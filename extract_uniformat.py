#!/usr/bin/env python3
import sys

try:
    import pandas as pd
    
    # Read the Excel file
    file_path = '/Users/josephagunbiade/Desktop/studio/onyx/Uniformat Building Elements Code.xlsx'
    
    # Try to read all sheets
    xlsx = pd.ExcelFile(file_path)
    print(f'Sheet names: {xlsx.sheet_names}')
    print('\n' + '='*80 + '\n')
    
    # Read each sheet
    for sheet_name in xlsx.sheet_names:
        print(f'Sheet: {sheet_name}')
        print('-'*40)
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(df.to_string())
        print('\n' + '='*80 + '\n')
        
except ImportError:
    print("pandas not installed. Trying openpyxl...")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook('/Users/josephagunbiade/Desktop/studio/onyx/Uniformat Building Elements Code.xlsx', read_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f'\nSheet: {sheet_name}')
            print('-'*40)
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    print('\t'.join(str(cell) if cell is not None else '' for cell in row))
                    
    except ImportError:
        print("Neither pandas nor openpyxl is installed.")
        print("Trying to use csv if possible...")
        
        import csv
        import zipfile
        
        # Excel files are zip archives, try to extract data
        try:
            with zipfile.ZipFile('/Users/josephagunbiade/Desktop/studio/onyx/Uniformat Building Elements Code.xlsx', 'r') as z:
                print("Files in Excel archive:")
                for name in z.namelist():
                    print(f"  {name}")
                    if 'sharedStrings.xml' in name or 'sheet1.xml' in name:
                        print(f"\nContent of {name}:")
                        print(z.read(name).decode('utf-8')[:1000])  # First 1000 chars
        except Exception as e:
            print(f"Error reading as zip: {e}")
            
except Exception as e:
    print(f'Error: {e}')